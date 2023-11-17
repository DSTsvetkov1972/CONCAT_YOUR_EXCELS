import openpyxl
import os






wb = openpyxl.load_workbook(os.path.join('Исходники','2022','Комментарии Зарубеж_01_2022.xlsx'))
sheet = wb.worksheets[0]
print(dir(sheet))
print(sheet.title)
row_count = sheet.max_row
column_count = sheet.max_column
print(sheet.max_row)
print(sheet.max_column)