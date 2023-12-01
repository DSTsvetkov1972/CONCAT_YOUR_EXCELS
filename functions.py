import pandas as pd
import numpy as np
import openpyxl
import os 
import csv
import win32com.client
from tkinter import *
from tkinter import messagebox
import tkinter as tk
from datetime import datetime
from colorama import init, Fore, Back, Style
from service_functions import *

import warnings
warnings.filterwarnings('ignore')


def check_files_isnt_open (TITLE):
    for file in ['.sheets.xlsx','.headers.xlsx','.statistics.xlsx']:
        if os.path.exists(os.path.join(os.getcwd(),'~$' + file)):
            sw_message = f'Закройте таблицу {file} и повторите попытку!'
            print(Fore.YELLOW + sw_message.replace('\n', ' ') + Fore.WHITE)
            messagebox.showwarning(TITLE,sw_message)
            return  False
    return True

def check_books (concat_tables_button):
    """
    Проверяем не были ли изменения в файлах в папке Исходники
    """
    source_folder = os.walk('Исходники')
    sheets_list = []
    if os.path.exists('.sheets.csv'):   
        sheets_in_csv_df = pd.read_csv('.sheets.csv', sep ='\t')[['Папка', 'Книга','Последнее обновление исходника']]
    else:
        sheets_in_csv_df = pd.DataFrame(columns=['Папка', 'Книга','Последнее обновление исходника'])
    sheets_in_csv_df.drop_duplicates( subset = ['Папка','Книга','Последнее обновление исходника'], inplace= True) 
    sheets_in_csv_df['Последнее обновление исходника'] = sheets_in_csv_df['Последнее обновление исходника'].apply(int)
    #print(Fore.MAGENTA + '',sheets_in_csv_df)
    for i in source_folder:
        folder =i[0]
        files = i[2]
        for file in files:
            if '~' in file:
                continue
            elif file[-5:] in ['.xlsx','.xlsm']:
                sheets_list.append({'Папка':folder,'Книга':file,'Последнее обновление исходника':int(os.path.getmtime(os.path.join(os.getcwd(),folder,file)))})

    sheets_df = pd.DataFrame(sheets_list)
    #sheets_df['Есть в Исходники'] = True
    
    compare_df = pd.concat([sheets_in_csv_df,sheets_df])
    #print(Fore.CYAN + '', compare_df , Fore.WHITE)
    compare_df = compare_df.drop_duplicates()
    if len(compare_df) != len(sheets_in_csv_df) or len(compare_df) != len(sheets_df):
        messagebox.showwarning(TITLE,'Книги в папке "Исходники" были добавлены или удалены или измененены.\nПросмотрите листы, чтобы обновить информацию!')
        #get_headers_button.pack_forget()
        concat_tables_button.pack_forget()
        get_sheets()
        return False
    else:
        return True

def check_multiple_headers(concat_tables_button):
    try:
        headers_df = pd.read_csv('.headers.csv', sep ='\t')
    except pd.errors.EmptyDataError:
        print(Fore.RED + 'Обработка невозможна! Вероятно отсутствуют книги или листы для которых требуется собрать информацию?' + Fore.WHITE)
        return
    headers_df_gropby = headers_df.groupby(['Папка','Книга','Лист']).count().reset_index()
    headers_df_gropby = headers_df_gropby.rename(columns = {'Найден по колонке':'Повторов'})
    headers_df_gropby = headers_df_gropby[headers_df_gropby['Повторов'] > 1]
    headers_df_gropby = headers_df_gropby[['Папка','Книга','Лист']]
    multiple_headers_df = headers_df_gropby.merge(headers_df,
                                                   how='inner',
                                                   on = ['Папка','Книга','Лист'] )
    multiple_headers_df.to_csv('.multiple_headers.csv', sep = '\t', index=None)
    if len(multiple_headers_df) == 0:
        print(Fore.GREEN + 'Случаи когда в одной таблице обнаружены несколько заголовков отсутствуют!' + Fore.WHITE)
        return True
    else:
        concat_tables_button.pack_forget()
        print(Fore.RED + 'Есть случаи когда в одной таблице обнаружены несколько заголовков!\nДобавьте неправильные закголвки в таблицу на листе Exceptions!' + Fore.WHITE)
        return False

def check_no_header(tables_from_sheets_dict,sheets_for_processing_list): 
    """
    Возвращает истину если нет таблиц с необнаружеными заголовками
    """   
    try: 
        pd.read_csv('.table_with_not_located_headers.csv', sep = '\t')
        print(Fore.RED + 'Есть таблицы для которых не найдены заголовки!' + Fore.WHITE)
        return False
    except pd.errors.EmptyDataError:
        print(Fore.GREEN + 'Заголовки найдены для всех таблиц!' + Fore.WHITE)        
        return True

def check_identical_column_names(list_to_check):
    list_to_check = list_to_check
    identical_column_names_list = []
    while list_to_check:
        i = list_to_check[0]
        list_to_check = list_to_check[1:]
        if i in list_to_check and i not in identical_column_names_list:
            identical_column_names_list.append(i)
    return identical_column_names_list

def check_unmade_decisions(concat_tables_button):
    df = pd.read_excel('.sheets.xlsx')
    df = df.fillna('',inplace = False)
    #print(df)
    #print(df[(df['Добавить'] == '')])
    #print(df[df['Добавить'] == 'ДА'])
    #print(df[(df['Добавить'] == '') | (df['Добавить'] == 'ДА')])
    #print(len(df[(df['Добавить'] == '') | (df['Добавить'] == 'ДА')]) == len(df))
    if len(df[(df['Добавить'] == '') | (df['Добавить'] == 'ДА')]) != len(df):
        sw_message = 'Для некоторых листов не принято решение нужно ли их добавлять в итоговую таблицу!\nВведите в колонку "Добавить" значение "ДА" для листов которые нужно добавлять, для остальных - пустое значение!'
        print(Fore.YELLOW + sw_message.replace('\n',' '))
        messagebox.showwarning(TITLE,sw_message)
        get_sheets()
        return False
    elif len(df[df['Добавить'] == '']) == len(df):
        sw_message = 'Ни один из листов в книгах из папки Исходники не оторбран для добавления в итоговую таблицу!\nВведите в колонку "Добавить" значение "ДА" для листов которые нужно добавлять, для остальных - пустое значение!'
        print(Fore.YELLOW + sw_message.replace('\n',' '))
        messagebox.showwarning(TITLE,sw_message )
        concat_tables_button.pack_forget()
        get_sheets()
        return False       
    else:    
        return True

def proceed_type(proceed_type):
    def wrapper(func):
        func.proceed_type = proceed_type
        return func
    return wrapper

def start_finish_time(func):
    def wrapper(*args, **kwargs):
        start = datetime.now()
        print(Fore.BLUE + '\n', start,'\tОбработка %s стартовала...'%(func.proceed_type) + Fore.WHITE)
        value = func(*args, **kwargs)
        finish = datetime.now()
        print(Fore.CYAN, finish,'\tОбработка %s завершена!\n'%(func.proceed_type) + Fore.WHITE)
        return value
    return wrapper

def get_book_info(folder,file,sheets_in_csv_df,sheets_in_xlsx_df,sheets_list):
    print(Fore.WHITE + f'Папка: {folder} Книга: {file} сканируем листы...')    
    file_in_csv_info_df = sheets_in_csv_df[(sheets_in_csv_df['Папка'] == folder) &
                                            (sheets_in_csv_df['Книга']  == file)]#  &

    
    if  ((len(file_in_csv_info_df) > 0 and file_in_csv_info_df['Последнее обновление исходника'].iloc[0] != int(os.path.getmtime(os.path.join(os.getcwd(),folder,file)))) or 
         len(file_in_csv_info_df) == 0):

        wb = openpyxl.load_workbook(os.path.join(folder,file))
        sheets = wb.worksheets   
        for sheet in sheets: 
            sheet_in_xlsx_info_df = sheets_in_xlsx_df[(sheets_in_xlsx_df['Папка'] == folder) &
                                            (sheets_in_xlsx_df['Книга'] == file)  & 
                                            (sheets_in_xlsx_df['Лист']  == sheet.title)]
            #print(Fore.MAGENTA + '\n', folder, file, sheet, sep = '\n')
            #print(Fore.RED + '', sheet_in_xlsx_info_df)
            if  len(sheet_in_xlsx_info_df ) > 0 :
                sheets_list.append([folder,
                                    file,
                                    sheet.title,
                                    sheet.max_row,
                                    sheet.max_column,
                                    sheet_in_xlsx_info_df['Добавить'].iloc[0],
                                    sheet_in_xlsx_info_df['Сколько строк нужно'].iloc[0],
                                    sheet_in_xlsx_info_df['Сколько колонок нужно'].iloc[0],
                                    int(os.path.getmtime(os.path.join(os.getcwd(),folder,file)))
                                    ])
            else:    
                sheets_list.append([folder,
                                    file,
                                    sheet.title,
                                    sheet.max_row,
                                    sheet.max_column,
                                    'ДА - если нужно обработать',
                                    sheet.max_row,
                                    sheet.max_column,
                                    int(os.path.getmtime(os.path.join(os.getcwd(),folder,file)))
                                    ])
            #print(Fore.BLACK + f'Папка: {folder} Книга: {file}', end = ' ')                                
            print(Fore.GREEN + f'Лист: {sheet.title} Строк: {sheet.max_row} Колонок: {sheet.max_column}' + Fore.WHITE)
    
    else:
        for sheet in file_in_csv_info_df.itertuples():
            sheet_in_xlsx_info_df = sheets_in_xlsx_df[(sheets_in_xlsx_df['Папка'] == folder) &
                                (sheets_in_xlsx_df['Книга'] == file)  & 
                                (sheets_in_xlsx_df['Лист']  == sheet[3])]
            #print(sheet)
            #print(sheets_in_xlsx_df)
            #print(folder,file,sheet[3])
            #print(sheet_in_xlsx_info_df)
            #print(sheet_in_xlsx_info_df['Строк на листе'])
            sheets_list.append([folder,
                    file,
                    sheet[3],
                    sheet_in_xlsx_info_df['Строк на листе'].iloc[0],
                    sheet_in_xlsx_info_df['Столбцов на листе'].iloc[0],
                    sheet_in_xlsx_info_df['Добавить'].iloc[0],
                    sheet_in_xlsx_info_df['Сколько строк нужно'].iloc[0],
                    sheet_in_xlsx_info_df['Сколько колонок нужно'].iloc[0],
                    int(os.path.getmtime(os.path.join(os.getcwd(),folder,file)))
                    ])
            #print(Fore.BLACK + f'Папка: {folder} Книга: {file}', end = ' ')                                
            print(Fore.GREEN + f'Лист: {sheet[3]} параметры уже были получены' + Fore.WHITE)
    
    #print(Fore.GREEN + '', folder, file,os.path.getmtime(os.path.join(os.getcwd(),folder,file)) ) 
    #print(Fore.GREEN + '', folder, sheet_in_csv_info_df['Последнее обновление исходника'].iloc[0] )     
    #print(Fore.RED +'', sheets_in_csv_df)  
    #print(Fore.CYAN + '', sheet_in_csv_info_df)
    #print(Fore.CYAN + '', sheet_info_df.columns)    

@start_finish_time
@proceed_type('"Сканирование листов в книгах"')
def get_sheets():
    """
    Функция получает список листов во всех экселевских книгах в папке Исходники
    и загружает его в файл .sheets.csv
    Затем функция открывает на рабочем столе файл .sheets.xlsx
    """
    if not check_files_isnt_open(TITLE): return 

    source_folder = os.walk('Исходники')
    sheets_list = []
    if os.path.exists('.sheets.csv'):   
        sheets_in_csv_df = pd.read_csv('.sheets.csv', sep ='\t')
    else:
        sheets_in_csv_df = pd.DataFrame(columns=['Папка', 'Книга','Последнее обновление исходника'])
    sheets_in_xlsx_df = pd.read_excel('.sheets.xlsx', header = 0)
    #print(sheets_in_xlsx_df)


    for i in source_folder:
        folder =i[0]
        files = i[2]
        for file in files:
            if '~' in file:
                continue
            elif file[-5:] in ['.xlsx','.xlsm']:
                get_book_info(folder,file,sheets_in_csv_df,sheets_in_xlsx_df,sheets_list)
 
    with open('.sheets.csv', 'w', newline='', encoding='utf-8') as sheets_csv:
        writer = csv.writer(sheets_csv, delimiter='\t')
        writer.writerow(['Папка',
                         'Книга',
                         'Лист',
                         'Строк на листе',
                         'Столбцов на листе',
                         'Добавить',
                         'Сколько строк нужно',
                         'Сколько колонок нужно',
                         'Последнее обновление исходника'
                         ])                

    #print(Fore.RED +'', sheets_list, Fore.WHITE + '' )
    sheets_csv_df = pd.DataFrame(sheets_list, columns =['Папка','Книга','Лист','Строк на листе','Столбцов на листе','Добавить','Сколько строк нужно','Сколько колонок нужно','Последнее обновление исходника'])
    #print(Fore.MAGENTA +'', sheets_csv_df.columns)
    #sheets_csv_df.fillna('', how = 'all')
    
    sheets_csv_df.to_csv('.sheets.csv', index= False, encoding='utf-8', sep='\t')

    #os.system('start excel.exe %s'%('.sheets.xlsx'))

    fileName = os.path.join(os.getcwd(),'.sheets.xlsx')
    xl_get_sheets = win32com.client.DispatchEx("Excel.Application")
    wb_get_sheets = xl_get_sheets.Workbooks.Open(fileName)
    xl_get_sheets.Visible = True
    wb_get_sheets.RefreshAll()
    wb_get_sheets.Save()
    #wb_get_sheets.SaveAs(Filename=os.path.join(os.getcwd(),'.sheets.xlsx'))

    if sheets_list != []:
        pass
        #get_headers_button.pack(anchor = CENTER, pady = (25,0), padx=(0,0))

@start_finish_time
@proceed_type('"Создание списка таблиц"')
def get_tables_from_sheets(tables_from_sheets_dict, sheets_for_processing_list):
    #print('tables_from_sheets_dict_ДО\n',tables_from_sheets_dict.keys())

    # если поменялись отобранные листы, то проверяем какие таблицы нужно загрузить
    sheets_for_processing_df = pd.read_excel(os.path.join(os.getcwd(),'.sheets.xlsx'), header = 0)
    sheets_for_processing_df = sheets_for_processing_df[sheets_for_processing_df['Добавить'] == 'ДА']
    sheets_for_processing_list.clear()
    for row in sheets_for_processing_df.itertuples():
        sheets_for_processing_list.append(row[1:4]+row[7:])
   
    if os.path.exists(os.path.join(os.getcwd(),'.selected_sheets.csv')):
        sheets_for_processing_df_before = pd.read_csv(os.path.join(os.getcwd(),'.selected_sheets.csv'),sep='\t')
    else:
        sheets_for_processing_df.to_csv(os.path.join(os.getcwd(),'.selected_sheets.csv'),sep='\t',index = False)
        sheets_for_processing_df_before = sheets_for_processing_df
    #print('\nsheets_for_processing_df_before\n',sheets_for_processing_df_before)
    #print('\nsheets_for_processing_df\n',       sheets_for_processing_df)
    #print('\nsheets_for_processing_df_before.equals(sheets_for_processing_df)',sheets_for_processing_df_before.equals(sheets_for_processing_df))
    #print('\ntables_from_sheets_dict != ()\n',tables_from_sheets_dict)
    if sheets_for_processing_df_before.equals(sheets_for_processing_df) and tables_from_sheets_dict != {}: 
        print(Fore.GREEN + 'Список листов, которые нужно загрузить не менялся, таблицы уже загружены!' + Fore.WHITE)
    else:        
        for row in sheets_for_processing_df.iterrows():
            #try:
                folder        = row[1]['Папка']
                book          = row[1]['Книга']
                sheet         = row[1]['Лист']
                rows_limit    = int(row[1]['Сколько строк нужно'])
                columns_limit = int(row[1]['Сколько колонок нужно'])
                print(Fore.WHITE + f'Книга: {folder} Папка: {book} Лист: {sheet}', end =' ')
                # Таблицу загружаем только если ранее на загружали или если файл изменился
                if  (((folder,book,sheet,rows_limit,columns_limit) not in tables_from_sheets_dict) or 
                     (((folder,book,sheet,rows_limit,columns_limit) in tables_from_sheets_dict) and tables_from_sheets_dict[(folder,book,sheet,rows_limit,columns_limit)]['Последнее обновление исходника'] != int(os.path.getmtime(os.path.join(os.getcwd(),folder,book))))
                     ):
                    try:
                        table = pd.read_excel(os.path.join(folder,book), sheet_name = sheet, header = None,  nrows = rows_limit, usecols = range(columns_limit))#.iloc[:,:columns_limit]
                    except MemoryError:
                        error_message = f'Папка: {folder}\nКнига: {book}\nЛист: {sheet}\nТАБЛИЦА СЛИШКОМ ВЕЛИКА И НЕ МОЖЕТ БЫТЬ ОБРАБОТАНА!!!\nПопробуйте ограничить количество строк и/или колонок, которые нужны из этой таблицы!'
                        print(Fore.RED + ' - ТАБЛИЦА СЛИШКОМ ВЕЛИКА И НЕ МОЖЕТ БЫТЬ ОБРАБОТАНА!!! Попробуйте ограничить количество строк и/или колонок, которые нужны из этой таблицы!' + Fore.WHITE)
                        messagebox.showerror(TITLE,error_message.replace('\\n',''))
                        return False
                    except pd.errors.ParserError:
                            table = pd.read_excel(os.path.join(folder,book), sheet_name = sheet, header = None,  nrows = rows_limit)#.iloc[:,:columns_limit]    
                    except Exception as e:
                        error_message = f'Папка: {folder}\nКнига: {book}\nЛист: {sheet}\nПРОИЗОШЛО ЧТО-ТО НЕПРЕДВИДЕННОЕ (код ошибки {e})!!!\nОбратитесь к разработчикам!'
                        print(Fore.RED + error_message)
                        messagebox.showerror(TITLE,error_message.replace('\\n',''))
                        return False
                    table.insert(0, 'Папка' , folder)
                    table.insert(1, 'Книга' , book) 
                    table.insert(2, 'Лист'  , sheet)
                    table.insert(3, 'Строка в исходнике', table.index.tolist())
                    tables_from_sheets_dict[(folder,book,sheet,rows_limit,columns_limit)]={'Таблица': table,
                                                                                           'Превью': table.iloc[:30,:30],
                                                                                           'Размер датафрейма (кБ)': round(table.memory_usage(deep=True).sum()/1024,2),
                                                                                           'Последнее обновление исходника':int(os.path.getmtime(os.path.join(os.getcwd(),folder,book)))}
                    table_size = tables_from_sheets_dict[(folder,book,sheet,rows_limit,columns_limit)]['Размер датафрейма (кБ)']
                    print(Fore.GREEN + f' - загрузили из экселя ({table_size:,}кБ)')
                else:
                    table_size = tables_from_sheets_dict[(folder,book,sheet,rows_limit,columns_limit)]['Размер датафрейма (кБ)']
                    print(Fore.GREEN + f' - таблица с листа уже была загружена ({table_size:,}кБ)')
                
            #except:
        #print(Fore.RED, tables_from_sheets_dict, Fore.WHITE)

        sheets_for_processing_df.to_csv(os.path.join(os.getcwd(),'.selected_sheets.csv'),sep='\t',index= False)
    return True
   # print('tables_from_sheets_dict_ПОСЛЕ\n',tables_from_sheets_dict.keys())


def get_exceptions():
    exceptions_df = pd.read_excel('.headers.xlsx', sheet_name='Exceptions')
    return [list(i[1:]) for i in exceptions_df.itertuples()]

@start_finish_time
@proceed_type('"Получение заголовков"')
def get_headers(tables_from_sheets_dict, sheets_for_processing_list):

    exceptions_list = get_exceptions()
    if not get_tables_from_sheets(tables_from_sheets_dict,sheets_for_processing_list):
        return False
    headers_specifications_df = pd.read_excel('.headers.xlsx', sheet_name = 'Settings')  #print(headers_specifications_df)  

    all_tables_headers_df = pd.DataFrame()
    table_with_not_located_headers = pd.DataFrame()
    
    for sheet_for_processing_list in sheets_for_processing_list:
        print(Fore.WHITE + "Папка: {} Книга: {} Лист: {}".format(*sheet_for_processing_list), end = ' ')
        table_headers_df = pd.DataFrame()
        
        for specification in headers_specifications_df.iterrows():
            column_number = specification[1]['Колонка']
            sign          = specification[1]['Признак']
            #print('Колонка: %s Признак: %s'%(column_number,sign))  
            #print('-'*50) 

            # отрабатываем то, что в таблице может быть меньше колонок чем в спецификации на поиск заголовка
            try:          
                header_df = tables_from_sheets_dict[(sheet_for_processing_list)]['Превью'][tables_from_sheets_dict[(sheet_for_processing_list)]['Превью'][column_number] == sign ] 
            except KeyError:
                header_df = pd.DataFrame() 

            #print(header_df)
            #print('-'*50)
            if len(header_df) > 0:
                #print('header_df\n', header_df ) 
                header_list = list(header_df.iloc[0])[:4]
                header_list.append(column_number)
                header_list.append(sign)
                #print('header_list \n', header_list ) 
                header_df.insert(4, 'Найден по колонке' , column_number)
                header_df.insert(5, 'По признаку' , sign)
                header_df.fillna('')
                #table_from_sheets_list['Строка заголовка']  = (header_df['Строка'].iloc[0])
                #print(Fore.MAGENTA,exceptions_list)
                #print(Fore.MAGENTA,header_list)
                if header_list not in exceptions_list:
                    table_headers_df = table_headers_df._append(header_df)
                    #print(table_headers_df)
                    tables_from_sheets_dict[(sheet_for_processing_list)]['Строка заголовка']  = table_headers_df.index[0]
                    #print(table_from_sheets_list['Строка заголовка'] )
                else:
                    print(Fore.YELLOW + ' - есть заголовок из списка исключений' + Fore.WHITE, end = ' ')
                
                #input()
            #print('table_headers_df\n',table_headers_df)
            #print('$'*50)    
        if len(table_headers_df) == 0:
            table_headers_df = pd.DataFrame([{'Папка':sheet_for_processing_list[0],
                                              'Книга':sheet_for_processing_list[1],
                                              'Лист':sheet_for_processing_list[2],
                                              #!!!!!!!!!!!!!!!!!!!!!!!
                                              'Строка в исходнике':'',
                                              'Найден по колонке':'',
                                              'По признаку':''
                                              }])
            table_with_not_located_headers = tables_from_sheets_dict[(sheet_for_processing_list)]['Превью']
            tables_from_sheets_dict[(sheet_for_processing_list)]['Строка заголовка']  = None  
            print(Fore.RED + ' - заголовки не найдены')
        else:
            print(Fore.GREEN + ' - заголовки найдены')
        
        table_with_not_located_headers.to_csv('.table_with_not_located_headers.csv', sep = '\t', index= False)   
        all_tables_headers_df = all_tables_headers_df._append(table_headers_df, ignore_index = True)
        
    all_tables_headers_df.to_csv('.headers.csv', sep = '\t', index= False)
    return True


 
#@start_finish_time
#@proceed_type('"Отобразить .headers.xlsx"')
def open_headers_xls(tables_from_sheets_dict,sheets_for_processing_list,concat_tables_button):

    """
    Функция открывает файл .headers.xlsx
    """
    concat_tables_button.pack_forget()
    if not check_files_isnt_open (TITLE): return
        
    # Проверяем не менялись ли исходники
    if not check_books (concat_tables_button): return

    # Проверяем для всех листов заполнено "ДА" или ПУСТО в колонке "Добавить" книги .sheets.xlsx
    if not check_unmade_decisions(concat_tables_button): return

    """
    if len(sheets_for_processing_df) != len(sheets_for_processing_df [(sheets_for_processing_df ['Добавить'] == 'ДА') | (sheets_for_processing_df ['Добавить'] =='')]):
        get_headers_button.pack_forget()
        concat_tables_button.pack_forget()
        messagebox.showwarning(TITLE, 'Для некоторых листов непринято решение нужно ли их добавлять или нет в итоговую таблцы.\nКолонка "Добавить" должна быть пустой или иметь значение "ДА"')
        return
    """

    get_headers(tables_from_sheets_dict, sheets_for_processing_list)
 
    if check_multiple_headers(concat_tables_button) and check_no_header(tables_from_sheets_dict,sheets_for_processing_list):
        fileName = os.path.join(os.getcwd(),'.headers.xlsx')
        xl = win32com.client.DispatchEx("Excel.Application")
        wb = xl.Workbooks.Open(fileName)
        xl.Visible = True
        wb.RefreshAll() 
        concat_tables_button.pack(anchor = CENTER, pady=(25,0))         
        return True
    else:
        fileName = os.path.join(os.getcwd(),'.headers.xlsx')
        xl = win32com.client.DispatchEx("Excel.Application")
        wb = xl.Workbooks.Open(fileName)
        xl.Visible = True
        wb.RefreshAll() 
        return False
        
@start_finish_time
@proceed_type('"Объединение таблиц"')
def concat_tables(tables_from_sheets_dict, sheets_for_processing_list, concat_tables_button):
    # Проверяем не открыты ли файлы, которые могут быть открыты в ходе выполения скрипта   
    if not check_files_isnt_open(TITLE): return

    # Проверяем не менялись ли исходники
    if not check_books (concat_tables_button): return

    # Проверяем для всех листов заполнено "ДА" или ПУСТО в колонке "Добавить" книги .sheets.xlsx
    if not check_unmade_decisions(concat_tables_button): return

    #if not open_headers_xls(tables_from_sheets_dict,sheets_for_processing_list): return 

    sheets_for_processing_df = pd.read_excel(os.path.join(os.getcwd(),'.sheets.xlsx'), header = 0)
    sheets_for_processing_df.fillna('', inplace=True)
    sheets_for_processing_df = sheets_for_processing_df[sheets_for_processing_df['Добавить'] == 'ДА']
    sheets_for_processing_list_actual = []
    sheets_for_processing_list_cant_add = []

    # Проверяем не менялся ли список листов которые нужно собрать  
    for row in sheets_for_processing_df.itertuples():
        sheets_for_processing_list_actual.append(row[1:4]+row[7:])
    if  sorted(sheets_for_processing_list_actual) != sorted(sheets_for_processing_list):
        #concat_tables_button.pack_forget()
        #open_headers_xls(tables_from_sheets_dict,sheets_for_processing_list,concat_tables_button)
        sw_message = "Изменился список листов, таблицы с которых нужно объеденить.\nТребуется пересобрать заголовки!"
        print(Fore.YELLOW + sw_message.replace('\n', ' ') + Fore.WHITE)
        messagebox.showwarning(TITLE, sw_message)  
        concat_tables_button.pack_forget()
        return

        

    
    if not check_no_header(tables_from_sheets_dict,sheets_for_processing_list) or not check_multiple_headers(concat_tables_button):
       open_headers_xls(tables_from_sheets_dict,sheets_for_processing_list,concat_tables_button)
       return
    
    total_table_df = pd.DataFrame()
    for sheet_for_processing_list in sheets_for_processing_list:
            #print(Fore.RED+'', sheet_for_processing_list, Fore.WHITE+'')
            print(Fore.WHITE + 'Папка: {} Книга: {} Лист: {}'.format(*sheet_for_processing_list), end = ' ')
        #try: 
            header_row = tables_from_sheets_dict[(sheet_for_processing_list)]['Строка заголовка']
            header_list = list(tables_from_sheets_dict[(sheet_for_processing_list)]['Таблица'].iloc[header_row])[4:]
            column_names_raw = ['Папка','Книга','Лист','Строка в исходнике'] + header_list
            i=0
            column_names = []
            for column_name_raw in column_names_raw:
                if pd.isna(column_name_raw):
                    column_name = '_' + str(i)
                    i += 1
                else:
                    column_name = column_name_raw
                column_names.append(column_name)
            identical_column_names = ', '.join(list(map(str, check_identical_column_names(column_names))))
            if identical_column_names:
                sheet_for_processing_list_cant_add = list(sheet_for_processing_list[:3])#.copy()
                sheet_for_processing_list_cant_add.append('Колонки: ' + identical_column_names + ' встречаются более одного раза!')
                sheets_for_processing_list_cant_add.append(sheet_for_processing_list_cant_add)
                print(Fore.RED + ' - таблица не может быть обработана так как названия колонок ',
                      Fore.WHITE + identical_column_names,
                      Fore.RED + ' встречаются более одного раза!' + Fore.WHITE)
                continue
            result_table = tables_from_sheets_dict[(sheet_for_processing_list)]['Таблица'][header_row+1:]
            result_table.columns = column_names
            result_table = result_table.dropna(how='all', axis = 1)
            total_table_df = total_table_df._append(result_table, ignore_index = True)
            print(Fore.GREEN + ' - данные извлечены и добавлены в итоговую таблицу' + Fore.WHITE)
        #except:
        #    sheets_for_processing_list_cant_add.append(sheets_for_processing_list_cant_add)
        #    print(Fore.RED + 'НЕ УДАЛОСЬ ИЗВЛЕЧЬ ДАННЫЕ ПО ПРИЧИНЕ РАНЕЕ НЕ ВСТРЕЧАВШЕЙСЯ! ОБРАТИТЕСЬ К РАЗРАБОТЧИКАМ!')
        #    continue
        #finally:
        #    print(total_table_df)
            
            #print(Fore.CYAN + 'total_table_df\n' + Fore.CYAN,total_table_df)
      
    total_table_df = total_table_df.replace('',np.nan)
    total_table_df_columns = total_table_df.columns[4:]
    total_table_df = total_table_df.dropna(axis=0,subset=total_table_df_columns,how='all')
    total_table_df = total_table_df.dropna(axis=1, how = 'all')
    total_table_df = total_table_df.fillna('')
    total_table_df = total_table_df.map(lambda x: str(x).replace(chr(10),''))#.replace(r'\n','\\n'))
    total_table_df.index.rename('Строка в итоговой таблице', inplace= True )


    print(Fore.BLUE + '', datetime.now(),'\t записываем результат в RESULT.csv' + Fore.WHITE)
    total_table_df.to_csv('RESULT.csv', sep ='\t')
    print(Fore.CYAN + '', datetime.now(),'\t результат записан в RESULT.csv' + Fore.WHITE)





    print(Fore.BLUE + '', datetime.now(),'\t проверяем правильно ли всё записалось' + Fore.WHITE) 
    #print(sheets_for_processing_list)
    #tables_size_list = [(sheet_for_processing_list[0],sheet_for_processing_list[1],sheet_for_processing_list[2],sheet_for_processing_list[3],sheet_for_processing_list[4],tables_from_sheets_dict[sheet_for_processing_list]['Размер датафрейма (кБ)']) for sheet_for_processing_list in sheets_for_processing_list]
    sheets_for_processing_list =  [(*sheet_for_processing_list,tables_from_sheets_dict[sheet_for_processing_list]['Размер датафрейма (кБ)']) for sheet_for_processing_list in sheets_for_processing_list]
    sheets_for_processing_df = pd.DataFrame(sheets_for_processing_list, columns = ['Папка','Книга','Лист','Строк для сканирования','Колонок для сканирования','Размер датафрейма (кБ)'])
    #print(tables_size_df)
    total_table_df_info = total_table_df.groupby(['Папка','Книга','Лист'])['Строка в исходнике'].agg('count').to_frame()   
    total_table_df_from_csv = pd.read_csv('RESULT.csv', sep ='\t')
    total_table_df_from_csv_info  = total_table_df_from_csv.groupby(['Папка','Книга','Лист'])['Строка в исходнике'].agg('count').to_frame()
    cant_add_df = pd.DataFrame(sheets_for_processing_list_cant_add, columns=['Папка','Книга','Лист','Комментарий'])

    compare_df = sheets_for_processing_df.merge(total_table_df_from_csv_info,
                                           on = ['Папка','Книга','Лист'],
                                           how = 'outer',
                                           suffixes=('_Итоговая таблица', '_Загружено в CSV'))
    compare_df = compare_df.merge(total_table_df_from_csv_info,
                                  on = ['Папка','Книга','Лист'],
                                  how = 'outer')
    compare_df = compare_df.fillna(0)
    compare_df = compare_df.merge(cant_add_df,
                                  on = ['Папка','Книга','Лист'],
                                  how = 'outer')
    compare_df = compare_df.fillna('')
    compare_df = compare_df.rename(columns={'Строка в исходнике_x':'В итоговой таблице после удаления пустых строк','Строка в исходнике_y':'Загружено в CSV'})                                            
    
    compare_df['Строк для сканирования']                         = compare_df['Строк для сканирования'].apply(int) 
    compare_df['В итоговой таблице после удаления пустых строк'] = compare_df['В итоговой таблице после удаления пустых строк'].apply(int)
    compare_df['Загружено в CSV']                                = compare_df['Загружено в CSV'].apply(int)
    compare_df.index.rename('№', inplace= True )
    #print(compare_df)                                           
    compare_df.to_csv('.statistics.csv')
    print(Fore.CYAN + '', datetime.now(),'\t проверка завершена' + Fore.WHITE)   
    
    fileName = os.path.join(os.getcwd(),'.statistics.xlsx')
    xl = win32com.client.DispatchEx("Excel.Application")
    wb = xl.Workbooks.Open(fileName)
    xl.Visible = True
    wb.RefreshAll()

    if len(sheets_for_processing_list_cant_add) > 0:
        print(Fore.RED + 'ВНИМАНИЕ: НЕКОТОРЫЕ ТАБЛИЦЫ НЕ УДАЛОСЬ ОБРАБОТАТЬ!' + Fore.WHITE)
        for sheet_for_processing_list_cant_add in sheets_for_processing_list_cant_add:
            print(Fore.RED + 'Папка: {} Книга: {} Лист: {} Комментарий: {}'.format(*sheet_for_processing_list_cant_add) + Fore.WHITE)
        sw_message = "Таблицы объеденены,\n*** НО НЕ ВСЕ ***!\nРезультат записан в RESULT.csv"
        print(Fore.YELLOW + sw_message.replace('\n', ' ') + Fore.WHITE)
        messagebox.showwarning(TITLE, sw_message)   
    else:
        si_message = "  Таблицы объеденены. Результат записан в RESULT.csv"
        print(Fore.MAGENTA + '-'*54,si_message.replace('\n', ' '),'-'*54 + Fore.WHITE, sep ='\n')
        messagebox.showinfo(TITLE, si_message)
    
    
    
    
    
    
    if not total_table_df_info.equals(total_table_df_from_csv_info):
        se_message = 'Таблицы объеденены, но почему-то в RESULT.csv записалось не то что насобиралось. Наверное исходные таблицы содержат недопустимые символы. Обратитесь к разработчикам для исправления ситуаци!'
        print(Fore.RED + '!'*50 + se_message.replace('\\n', ' ') + '!'*50 + Fore.WHITE)
        messagebox.showerror(TITLE, se_message)
    
    total_table_df_from_csv_len = len(open('RESULT.csv',encoding='utf-8').readlines())

    if (total_table_df_from_csv_len - 1) != len(total_table_df):
        se_message = 'Таблицы объеденены, но почему-то в RESULT.csv записалось не то что насобиралось. Наверное исходные таблицы содержат недопустимые символы. Обратитесь к разработчикам для исправления ситуаци!'
        print(Fore.RED + se_message.replace('\n', ' ') + Fore.WHITE)
        messagebox.showerror(TITLE, se_message)


    if not total_table_df_info.equals(total_table_df_from_csv_info):
        se_message = 'Таблицы объеденены, но почему-то агрегированные значения в RESULT.csv не совпадают с агрегированными значениями в total_table_df!'
        print(Fore.RED + se_message.replace('\\n', ' ') + Fore.WHITE)
        messagebox.showerror(TITLE, se_message)
